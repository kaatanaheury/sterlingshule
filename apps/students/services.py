"""Helpers for creating Student accounts and importing them in bulk."""
import csv
import io
from datetime import datetime

from django.contrib.auth.hashers import make_password
from django.db import transaction

from apps.academics.models import ClassLevel, Stream
from apps.accounts.models import Role, User

from .models import Student


def _unique_username(institution, base: str) -> str:
    """Return a username unique within (institution, username)."""
    base = (base or "student").strip().lower().replace(" ", "")
    candidate = base
    n = 1
    while User.objects.filter(institution=institution, username=candidate).exists():
        n += 1
        candidate = f"{base}{n}"
    return candidate


@transaction.atomic
def create_student_with_user(
    *,
    institution,
    first_name: str,
    second_name: str,
    third_name: str = "",
    gender: str = "",
    date_of_birth=None,
    class_level: ClassLevel,
    stream: Stream | None = None,
    admission_number: str,
    initial_password: str | None = None,
) -> Student:
    username = _unique_username(institution, first_name)
    password = (initial_password or "").strip() or admission_number or "student123"
    user = User.objects.create(
        username=username,
        password=make_password(password),
        role=Role.STUDENT,
        institution=institution,
        first_name=first_name,
        last_name=second_name,
        must_change_password=True,
    )
    student = Student.objects.create(
        institution=institution,
        user=user,
        first_name=first_name,
        second_name=second_name,
        third_name=third_name or "",
        gender=gender or "",
        date_of_birth=date_of_birth,
        class_level=class_level,
        stream=stream,
        admission_number=admission_number,
    )
    return student


def _parse_date(value):
    if not value:
        return None
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _read_rows(uploaded_file):
    """Yield dict rows from a CSV or XLSX upload."""
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        text = uploaded_file.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            yield {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
        return

    # Treat anything else (.xlsx, .xls) as Excel via openpyxl.
    from openpyxl import load_workbook

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    headers = [(c.value or "").strip().lower() if c.value else "" for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        record = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            record[h] = v if v is not None else ""
        yield record


def import_students(*, institution, uploaded_file) -> dict:
    """Import students from CSV/XLSX. Returns counts + per-row errors."""
    created = 0
    errors: list[str] = []

    class_levels = {
        c.name.lower(): c for c in ClassLevel.objects.filter(institution=institution)
    }
    streams_by_class: dict[tuple[int, str], Stream] = {
        (s.class_level_id, s.name.lower()): s
        for s in Stream.objects.filter(institution=institution)
    }

    for idx, raw in enumerate(_read_rows(uploaded_file), start=2):
        try:
            first_name = str(raw.get("first_name") or "").strip()
            second_name = str(raw.get("second_name") or "").strip()
            third_name = str(raw.get("third_name") or "").strip()
            adm = str(raw.get("admission_number") or "").strip()
            cl_name = str(raw.get("class_level") or "").strip().lower()
            stream_name = str(raw.get("stream") or "").strip().lower()
            gender = str(raw.get("gender") or "").strip().upper()[:1]
            dob = _parse_date(raw.get("date_of_birth"))

            if not (first_name and second_name and adm and cl_name):
                errors.append(
                    f"Row {idx}: missing required field(s) — first_name/second_name/admission_number/class_level."
                )
                continue
            if cl_name not in class_levels:
                errors.append(f"Row {idx}: unknown class_level '{raw.get('class_level')}'.")
                continue
            if Student.objects.filter(institution=institution, admission_number=adm).exists():
                errors.append(f"Row {idx}: admission_number '{adm}' already exists.")
                continue

            class_level = class_levels[cl_name]
            stream = streams_by_class.get((class_level.id, stream_name)) if stream_name else None

            create_student_with_user(
                institution=institution,
                first_name=first_name,
                second_name=second_name,
                third_name=third_name,
                gender=gender if gender in ("M", "F", "O") else "",
                date_of_birth=dob,
                class_level=class_level,
                stream=stream,
                admission_number=adm,
            )
            created += 1
        except Exception as exc:  # pragma: no cover — defensive
            errors.append(f"Row {idx}: {exc}")

    return {"created": created, "errors": errors}
